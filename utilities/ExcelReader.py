import openpyxl


import openpyxl

def get_data_from_excel(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    data_list = []
    for row in range(2, sheet.max_row + 1):  # start from 2 to skip the header
        email = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        data_list.append((email, password))
    return data_list